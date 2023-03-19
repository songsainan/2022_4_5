import openpyxl


# 定义一个excel操作类，创建对象时需传入文件名（name）和表单名（sheet）
class HandleExcel:
    def __init__(self, name, sheet):
        self.name = name
        self.sheet = sheet

    # 创建一个excel操作方法
    def read_excel(self):
        # 获取工作簿
        wb = openpyxl.load_workbook(self.name)
        # 获取表单
        sh = wb[self.sheet]
        # 按行读取表单全部内容，并转化为列表形式
        content = list(sh.rows)

        # 通过列表推导式获取表单首行内容
        title = [i.value for i in content[0]]

        # 创建空列表，用于存放表单首行之下的全部内容（即编写的用例数据）
        case = []

        # 通过for循环遍历表单首行之外内容，存于i
        for i in content[1:]:
            # 遍历i，并将每次遍历的值存为列表（参考上面列表推导式）
            data = [item.value for item in i]
            # 将首行内容[title]和其他每行内容[data]打包成一个字典存在列表case里
            dic = dict(zip(title, data))
            case.append(dic)

        # 最后返回case,后面才能调用
        return case

    def write_excel(self, row, column, value):
        # 获取工作簿
        wb = openpyxl.load_workbook(self.name)
        # 获取表单
        sh = wb[self.sheet]
        # 写入表格对象
        sh.cell(row=row, column=column, value=value)
        # 保存工作簿
        wb.save(self.name)


if __name__ == '__main__':
    excel = HandleExcel(r'C:\Users\27000\PycharmProjects\2022_4_5\cases\case.xlsx', 'recharge')
    print(excel.read_excel())
